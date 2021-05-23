import discord
from discord.ext import commands
import asyncio
import random
from openpyxl import load_workbook


class account:
    def random_email():
        wb = load_workbook("assets/accounts.xlsx")
        ws = wb.get_sheet_by_name('accounts')
        columnA = ws['A']
        email = [columnA[x].value for x in range(len(columnA))]
        random_email = random.choice(email)
        while random_email == None:
            random_email = random.choice(email)
        return random_email

    def random_password():
        wb = load_workbook("assets/accounts.xlsx")
        ws = wb.get_sheet_by_name('accounts')
        columnB = ws['B']
        password = [columnB[x].value for x in range(len(columnB))]
        random_password = str(random.choice(password)) + \
            str(random.choice(password))
        return random_password


async def generate(ctx, emd):
    await ctx.author.send(embed=emd)
    await ctx.send(embed=discord.Embed(
        description=f'{ctx.author.mention} **Check your DMs!**',
        colour=discord.Colour.random()
    ))
    await asyncio.sleep(20)
    await ctx.channel.purge(limit=3)


class gen(commands.Cog):

    def __init__(self, bot):
        self.bot = bot

    @commands.Cog.listener()
    async def on_ready(self):
        print('Załadowano gen')

    @commands.command()
    async def psc(self, ctx):

        digit = ['0', '1', '2', '3', '4', '5', '6', '7', '8', '9']

        emd = discord.Embed(
            title='Your PaySafeCard code is down here',
            description=f'||0{random.choice(digit)}{random.choice(digit)}{random.choice(digit)}-{random.choice(digit)}{random.choice(digit)}{random.choice(digit)}{random.choice(digit)}-{random.choice(digit)}{random.choice(digit)}{random.choice(digit)}{random.choice(digit)}-{random.choice(digit)}{random.choice(digit)}{random.choice(digit)}{random.choice(digit)}||',
            colour=discord.Colour.blue()
        )

        emd.set_footer(
            text='Generated via complicated algorithm by paysafecardgeneratorAPI.org')
        emd.set_thumbnail(
            url='https://avatars.githubusercontent.com/u/22189864?s=280&amp;v=4')

        await generate(ctx, emd)

    @commands.command()
    async def nitro(self, ctx):
        letters = ['0', '1', '2', '3', '4', '5', '6', '7', '8', '9', 'a', 'b', 'c', 'd', 'e', 'f', 'g', 'h', 'i', 'j', 'k', 'l', 'm', 'n', 'o', 'p', 'q', 'r', 's', 't',
                   'u', 'v', 'w', 'x', 'y', 'z', 'A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z']

        emd = discord.Embed(
            title='Your NITRO link is down here',
            description=f'||https://discord.gift/{random.choice(letters)}{random.choice(letters)}{random.choice(letters)}{random.choice(letters)}{random.choice(letters)}{random.choice(letters)}{random.choice(letters)}{random.choice(letters)}{random.choice(letters)}{random.choice(letters)}{random.choice(letters)}{random.choice(letters)}{random.choice(letters)}{random.choice(letters)}{random.choice(letters)}{random.choice(letters)}||',
            colour=discord.Colour.orange()
        )

        emd.set_thumbnail(url='https://i.redd.it/mvoen8wq3w831.png')

        await generate(ctx, emd)

    @commands.command()
    async def origin(self, ctx):
        emd = discord.Embed(
            title='Your Origin account password and email is down here',
            description=f'||Email: {account.random_email()}||\n||Password: {account.random_password()}||',
            colour=discord.Colour.orange()
        )
        emd.set_thumbnail(url='https://i.redd.it/mvoen8wq3w831.png')

        await generate(ctx, emd)


def setup(bot):
    bot.add_cog(gen(bot))
